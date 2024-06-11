from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import argparse
import os

usage = """
Usage:
This script takes a directory of tag pileup composite output files and writes an Frank-formatted excel document with composite data.

Example: python make_excel_composite.py -i /path/to/composites -o COMPOSITES.xlsx
"""

def getParams():
	'''Parse parameters from the command line'''
	parser = argparse.ArgumentParser(description='This script takes a directory of composite (*.out) files and combines them into an excel spreadsheet.')
	parser.add_argument('-i','--input', metavar='composite-dir', required=True, help='directory with all the composite data files')
	parser.add_argument('-o','--output', metavar='outfile', required=True, help='output name to save workbook to')

	args = parser.parse_args()
	return(args)

def load_composite(composite_file):
	'''Load arrays of scores'''
	all_composites = {}
	reader = open(composite_file,'r')
	# Update sequence with each variant
	for line in reader:
		if(line.find("\t")==0):
			continue
		tokens = line.strip().split("\t")
		if(tokens[0]=="NAME"):
			continue
		fn_tokens = tokens[0].split("_")
		key = "_".join(fn_tokens[:-2])
		all_composites.setdefault(key,([],[]))
		if(fn_tokens[-2]=="sense"):
			all_composites[key] = (tokens,all_composites[key][1])
		elif(fn_tokens[-2]=="anti"):
			all_composites[key] = (all_composites[key][0],tokens)
		else:
			all_composites[key] = (tokens,tokens)
	reader.close()
	return all_composites.values()

if __name__ == "__main__":
	'''Write BED file in new order.'''
	args = getParams()

	wb = Workbook()
	ws = wb.active

	## SET EXCEL FORMAT CONSTANTS
	ws.title = "CompositeData"
	FILL_BLACK = PatternFill("solid", fgColor="000000")
	FILL_BLUE = PatternFill("solid", fgColor="0000FF")
	FILL_RED = PatternFill("solid", fgColor="FF0000")
	FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
	FILL_GRAY = PatternFill("solid", fgColor="BFBFBF")
	THIN = Side(border_style="thin",color="000000")
	THICK = Side(border_style="thick",color="000000")

	## BUILD TEMPLATE
	ws['A1'] = "PEGR ID"
	ws['A2'] = "Target"
	ws['A3'] = "Ab"
	ws['A4'] = "Strain or Genotype"
	ws['A5'] = "Mutation"
	ws['A6'] = "Media"
	ws['A7'] = "Condition 1"
	ws['A8'] = "Condition 2"
	for cell in ws['A1':'A8']:
		cell[0].alignment = Alignment(horizontal='right')

	# Fill yellow sense title bar
	for cell in ws['B10':'KS10'][0]:
		cell.fill = FILL_YELLOW
		cell.border = Border(top=THIN, bottom=THIN, right=THIN, left=THIN)
	# Fill yellow anti title bar
	for cell in ws['B1110':'KS1110'][0]:
		cell.fill = FILL_YELLOW
		cell.border = Border(top=THIN, bottom=THIN, right=THIN, left=THIN)

	# Fill black middle zone
	ws['A10'].fill = FILL_BLACK
	for col in ws.iter_cols(min_row=1012, max_col=305, max_row=1108):
		for cell in col:
			cell.fill = FILL_BLACK
	for cell in ws['A2111':'KS2111'][0]:
		cell.fill = FILL_BLACK
	# for col in ws.iter_cols(min_row=2111, max_col=305, max_row=2111):
	# 	for cell in col:
	# 		cell.fill = FILL_BLACK

	# Fill blue for sense bordering cells
	for col in ws.iter_cols(min_row=1011, max_col=305, max_row=1011):
		for cell in col:
			cell.fill = FILL_BLUE
	# Fill red for anti bordering cells
	for col in ws.iter_cols(min_row=1109, max_col=305, max_row=1109):
		for cell in col:
			cell.fill = FILL_RED

	# Fill gray for last two rows
	for col in ws.iter_cols(min_row=1009, min_col=2, max_col=305, max_row=1010):
		for cell in col:
			cell.fill = FILL_GRAY
	for col in ws.iter_cols(min_row=2109, min_col=2, max_col=305, max_row=2110):
		for cell in col:
			cell.fill = FILL_GRAY

	# Fill and merge Sense y-axis label
	ws['A11'] = "Same (Sense)"
	ws['A11'].fill = FILL_BLUE
	ws['A11'].font = Font(size=16, color='FFFFFF', bold=True)
	ws['A11'].alignment = Alignment(text_rotation=90, vertical='center')
	ws.merge_cells('A11:A1011')
	# Fill and merge Anti y-axis label
	ws['A1109'] = "Opposite (Anti)"
	ws['A1109'].fill = FILL_RED
	ws['A1109'].font = Font(size=16, color='FFFFFF', bold=True)
	ws['A1109'].alignment = Alignment(text_rotation=90, vertical='center')
	ws.merge_cells('A1109:A2110')

	# Fill xvals for both sense and anti
	x_vals = range(-499,501)
	for i in range(1000):
		sense_cell = 'B%i' % (11+i)
		ws[sense_cell] = x_vals[i]
		ws[sense_cell].border = Border(right=THICK, left=THICK)
		anti_cell = 'B%i' % (1111+i)
		ws[anti_cell] = x_vals[i]
		ws[anti_cell].border = Border(right=THICK, left=THICK)
	ws['B10'].border = Border(top=THICK, bottom=THICK, right=THICK, left=THICK)
	ws['B11'].border = Border(top=THICK, right=THICK, left=THICK)
	ws['B1010'].border = Border(bottom=THICK, right=THICK, left=THICK)
	ws['B1110'].border = Border(top=THICK, bottom=THICK, right=THICK, left=THICK)
	ws['B1111'].border = Border(top=THICK, right=THICK, left=THICK)
	ws['B2110'].border = Border(bottom=THICK, right=THICK, left=THICK)

	# Hide Sense rows
	for idx in range(13, 1009):
		ws.row_dimensions[idx].hidden = True
	# Hide Anti rows
	for idx in range(1113, 2109):
		ws.row_dimensions[idx].hidden = True
	# Hide Black zone rows
	for idx in range(1013, 1106):
		ws.row_dimensions[idx].hidden = True

	## DYNAMIC FILL VALUES
	cidx = 3
	for filename in sorted(os.listdir(args.input)):
		if(os.path.splitext(filename)[1]!=".out"):
			continue
		# Parse metadata out of filename
		fn_tokens = filename.split("_")
		conditions = fn_tokens[6].split(",")
		if(len(conditions)>2):
			conditions[1] += "..check for third condition"
		elif(len(conditions)==1):
			conditions.append("-")
		_ = ws.cell(column=cidx, row=1, value=fn_tokens[0])
		_ = ws.cell(column=cidx, row=2, value=fn_tokens[1])
		_ = ws.cell(column=cidx, row=3, value=fn_tokens[2])
		_ = ws.cell(column=cidx, row=4, value=fn_tokens[3])
		_ = ws.cell(column=cidx, row=5, value=fn_tokens[4])
		_ = ws.cell(column=cidx, row=6, value=fn_tokens[5])
		_ = ws.cell(column=cidx, row=7, value=conditions[0])
		_ = ws.cell(column=cidx, row=8, value=conditions[1])
		# Load each file values
		composite_data = load_composite(args.input + "/" + filename)
		if(len(composite_data)>1):
			print("Warning: COMPOSITE DATA HAS MORE THAN ONE SAMPLE...")
			print(composite_data)
			quit()
		# Fill in excel workbook
		for data in composite_data:
			for i in range(len(data[0])):
				# Update sense
				_ = ws.cell(column=cidx, row=10+i, value=data[0][i])
				# Update anti
				_ = ws.cell(column=cidx, row=1110+i, value=data[1][i])
		cidx += 1
	wb.save(args.output)
